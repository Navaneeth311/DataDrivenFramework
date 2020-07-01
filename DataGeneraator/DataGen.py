import openpyxl


def dataGenerator():
    wb = openpyxl.load_workbook("/Users/navaneeth/Downloads/Flightz.xlsx")
    sh = wb['Sheet1']
    row = sh.max_row
    outer_list =[]
    inner_list = []
    for i in range(1, row+1):
        inner_list = []
        source = sh.cell(i,1)
        destination = sh.cell(i,2)
        inner_list.insert(0,source.value)
        inner_list.insert(1, destination.value)
        outer_list.insert(i-1, inner_list)
    return outer_list